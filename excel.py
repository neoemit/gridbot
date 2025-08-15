# gridbot - A Telegram bot for spreadsheet lookups
# Copyright (C) 2025  Thomas La Mendola
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from __future__ import annotations
import re
from pathlib import Path
from typing import Iterable
from openpyxl import load_workbook
import xlrd

HIDDEN_PREFIXES: tuple[str, ...] = (".", "~$")  # ignore macOS/Linux dotfiles and Office temp files

def list_excel_files(folder: Path) -> list[Path]:
    """Return sorted list of visible Excel files (.xls/.xlsx), ignoring hidden/temp files."""
    files: list[Path] = []
    if not folder.exists():
        return files
    for f in folder.iterdir():
        name = f.name
        if any(name.startswith(p) for p in HIDDEN_PREFIXES):
            continue
        if f.is_file() and f.suffix.lower() in (".xls", ".xlsx"):
            files.append(f)
    return sorted(files, key=lambda p: p.name.lower())

def list_sheets(file_path: Path) -> list[str]:
    if file_path.suffix.lower() == ".xlsx":
        wb = load_workbook(file_path, read_only=True)
        return list(wb.sheetnames)
    wb = xlrd.open_workbook(file_path.as_posix())
    return list(wb.sheet_names())

_CELL_RE = re.compile(r"^([A-Za-z]+)(\d+)$")

def _col_letters_to_index(letters: str) -> int:
    """Convert Excel column letters (e.g., 'C', 'AA') to 0-based index."""
    total = 0
    for ch in letters.upper():
        total = total * 26 + (ord(ch) - 64)
    return total - 1

def read_cell(file_path: Path, sheet_name: str, cell_coord: str):
    """Read a cell value. For .xlsx, try to resolve formulas using data_only; if empty, try xlcalculator."""
    if file_path.suffix.lower() == ".xlsx":
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[sheet_name]
        value = sheet[cell_coord].value
        if value is None:
            # Attempt to evaluate formula offline using xlcalculator
            try:
                from xlcalculator import ModelCompiler, Model
                from xlcalculator.reader import Reader
                reader = Reader(file_path.as_posix())
                compiler = ModelCompiler()
                model = compiler.read_and_parse(reader.read())
                calc = Model(model)
                value = calc.evaluate(f"{sheet_name}!{cell_coord}")
            except Exception as e:  # noqa: BLE001
                value = f"Error calculating formula: {e}"
        if isinstance(value, (int, float)):
            return f"R{value:,.2f}"
        return value

    # .xls via xlrd
    wb = xlrd.open_workbook(file_path.as_posix())
    sh = wb.sheet_by_name(sheet_name)
    m = _CELL_RE.match(cell_coord)
    if not m:
        return None
    col_letters, row_num = m.groups()
    col_idx = _col_letters_to_index(col_letters)
    row_idx = int(row_num) - 1
    value = sh.cell_value(row_idx, col_idx)
    if isinstance(value, (int, float)):
        return f"R{value:,.2f}"
    return value
